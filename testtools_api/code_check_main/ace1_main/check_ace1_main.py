import os, json, shutil
import re
import time
import openpyxl
from ace1whiteList import sourcs_parse
from get_source import get_source_json
from ext5 import produce_html_main
from ACEshell import write_total


def search_all(string, target):
    """
    寻找所有target字符串, 为了先大致筛选api content
    :param string:
    :param target:
    :return:
    """
    result_list = []
    while target in string:
        r = string.find(target)
        result_list.append(r)
        string = string.replace(target, len(target) * ' ', 1)
    return result_list


# xlsx追加写入内容
# value_list=[[],[],[]]会更快，减少了excel文件的打开保存
def write_excel_xlsx_append(path, sheet_name='summary', value_list=[[]]):
    # 判断文件是否存在
    if not os.path.exists(path):
        # 创建一个workbook 设置编码
        workbook = openpyxl.Workbook()
        # for name in sheet_name:
        #     workbook.create_sheet(name)
        workbook.create_sheet("summary", 0)
        workbook.remove(workbook['Sheet'])
        # 创建表summary
        summary = workbook['summary']
        title = ['归属子系统', '英文子系统', '模块名', '类名', '方法名*','方法名',
                 '使用次数','是否使用','是否误报','误报原因']
        summary.append(title)
        # 保存文件
        workbook.save(path)
        # print(f'文件{path}不存在，创建新表格')
    if len(value_list) != 0:
        workbook = openpyxl.load_workbook(path)
        for line in value_list:
            sheet = workbook[sheet_name]
            sheet.append(line)
        workbook.save(path)  # 保存工作簿
        # print(f"表格 {path}【追加】写入数据成功！")


# hml文件检索
def check_hml(target_dir):
    # 找到所有需要扫描胡target hml文件
    for root, dirs, files in os.walk(target_dir):
        for file in files:
            if file.endswith('.hml'):
                this_file_dir_name = os.path.basename(root)
                # 部分文件夹有子文件夹
                if this_file_dir_name in ['prop', 'router', 'style']:
                    this_file_dir_name = os.path.basename(os.path.dirname(root))

                # 匹配上之后需要检查
                if this_file_dir_name in all_source_json_list:
                    this_file = os.path.join(root, file)
                    this_source = os.path.join(TMP_DIR_PATH, this_file_dir_name)
                    with open(this_file, 'r', encoding='utf-8') as this_hml:
                        this_hml_content = this_hml.read()
                    with open(this_source, 'r', encoding='utf-8') as this_source_json:
                        this_source_dict = json.loads(this_source_json.read())

                    check_content_list = re.findall(f'(<{this_file_dir_name}.*?>)', this_hml_content, re.S)
                    for check_content in check_content_list:
                        parms = re.findall('([^" ]*)="', check_content)
                        for api in this_source_dict['api']:
                            for parm in parms:
                                if parm == api['api_method_name']:
                                    api['api_used_count'] += 1
                                    api['api_used'] = '是'
                    with open(this_source, 'w', encoding='utf-8') as this_source_json:
                        this_source_json.write(json.dumps(this_source_dict))



# css文件检索
def check_css(target_dir):
    source_css_json_path = os.path.join(TMP_DIR_PATH, 'cssConfig')
    with open(source_css_json_path, 'r') as source_css_json:
        source_css_dict = json.loads(source_css_json.read())
    # 找到所有需要扫描胡target css文件
    for root, dirs, files in os.walk(target_dir):
        for file in files:
            if file.endswith('.css'):
                this_file = os.path.join(root, file)
                with open(this_file, 'r', encoding='utf-8') as this_css:
                    this_css_content = this_css.read()
                all_css_parms = re.findall('([^\s]*):.*;', this_css_content)

                for api in source_css_dict['api']:
                    for parm in all_css_parms:
                        if parm == api['api_method_name']:
                            api['api_used_count'] += 1
                            api['api_used'] = '是'

    with open(source_css_json_path, 'w', encoding='utf-8') as this_source_json:
        this_source_json.write(json.dumps(source_css_dict))


def check_ace(source_path, target_path):
    global TMP_DIR_PATH, all_source_json_list

    totalapi_dict_list = []
    # 结果文件夹
    result_dir = f'./Ace1result_{time.time()}'
    save_result_xlsx_path = os.path.join(result_dir, 'ace1_0_result.xlsx')

    if os.path.exists(result_dir):
        shutil.rmtree(result_dir)
    os.makedirs(result_dir)

    TMP_DIR_PATH = os.path.join(result_dir, 'source_json')

    if os.path.exists(TMP_DIR_PATH):
        shutil.rmtree(TMP_DIR_PATH)
    os.makedirs(TMP_DIR_PATH)

    # 获取所有source基础数据
    all_source_json_list = get_source_json(source_path, TMP_DIR_PATH)

    check_hml(target_path)
    check_css(target_path)

    #####################
    # 生成结果数据写入excel
    check_index = 0
    for root, dirs, files in os.walk(TMP_DIR_PATH):
        for file in files:
            source_file_path = os.path.join(root, file)
            with open(source_file_path, 'r') as source_file:
                result_dict = json.loads(source_file.read())
                result_dict = sourcs_parse(result_dict)
                check_index += 1
                print(f'now check {source_file_path}, index: ', check_index)
                if not result_dict:
                    break
                for api in result_dict['api']:
                    totalapi_dict = {'subsystem_ch': 'ACE开发框架',
                                     'subsystem_en': 'ace',
                                     'api_module_name': result_dict['api_module_name'],
                                     'api_class_name': api['api_class_name'],
                                     'api_method_all':api['api_method_all'],
                                     'api_method_name': api['api_method_name'],
                                     'api_used_count': api['api_used_count'],
                                     'api_used':api['api_used'],
                                     'is_white':api['is_white'],
                                     'desc':api['desc']
                                     }

                    totalapi_dict_list.append(totalapi_dict)

    data_dict = totalapi_dict_list
    all_data = []
    for i in range(0, len(data_dict)):
        single_data = []
        single_data.append(data_dict[i]['subsystem_ch'])
        single_data.append(data_dict[i]['subsystem_en'])
        single_data.append(data_dict[i]['api_module_name'])
        single_data.append(data_dict[i]['api_class_name'])
        single_data.append(data_dict[i]['api_method_all'])
        single_data.append(data_dict[i]['api_method_name'])
        single_data.append(data_dict[i]['api_used_count'])
        single_data.append(data_dict[i]['api_used'])
        single_data.append(data_dict[i]['is_white'])
        single_data.append(data_dict[i]['desc'])
        all_data.append(single_data)

    write_excel_xlsx_append(save_result_xlsx_path, value_list=all_data)
    write_total(save_result_xlsx_path)
    print("正在生成页面信息")
    time.sleep(2)
    # 生成html结果文件
    produce_html_main(TMP_DIR_PATH, result_dir)
    print("页面生成完毕")
    print('扫描结果文件夹位置：' + result_dir)


if __name__ == '__main__':
    source_path = r'F:\work\NDK\API_Coverage_NDK_0221_test\interface_sdk-js-master\api\config'
    target_path = r'F:\work\NDK\API_Coverage_NDK_0221_test\xts_acts-master\ace\ace_ets_standard'
    check_ace(source_path, target_path)
