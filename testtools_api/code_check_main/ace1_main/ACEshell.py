import openpyxl
def write_total(path):
    # path = r"D:\wechar\WeChat Files\wxid_slguwwa9fk1w12\FileStorage\File\2022-03\0326\sdk_main3\SDKresult_1648265047.87342\sdk_result.xlsx"
    path = path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["summary"]
    rows = list(sheet.rows)
    system_lists = []
    for row in rows[1:]:
        system_lists.append(row[0].value)
        system_lists.append(row[6].value)

    system_sets = [i  for i in set(system_lists) if (type(i)!=int)]
    _len = len(system_lists)
    systems = []

    for system_set in system_sets:
        count = 0
        system = []
        all = system_lists.count(system_set)

        for system_index in range(0,_len,2):
            if system_set == system_lists[system_index]:
                if system_lists[system_index+1]>0:
                    count+=1
        system.append(system_set)
        system.append(all)
        system.append(count)
        system.append("%.2f%%" % (count/all*100))
        systems.append(system)
    system_alls = 0
    system_containt = 0
    for system_all in systems:
        system_alls += system_all[1]
        system_containt += system_all[2]
    systems.append(["全部",system_alls,system_containt,"%.2f%%" % (system_containt / system_alls * 100)])

    workbook.create_sheet("total", 1)
    total = workbook['total']
    title = ['中文子系统', '总API数量', '已经覆盖API', 'API覆盖率']
    total.append(title)
    workbook.save(path)
    if len(systems) != 0:
        workbook = openpyxl.load_workbook(path)
        for line in systems:
            sheet = workbook["total"]
            sheet.append(line)
        workbook.save(path)

if __name__ == '__main__':
    write_total(r"C:\Users\ts\PycharmProjects\dev07\utils\ndk.xlsx")


